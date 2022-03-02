# import openpyxl and tkinter modules
from openpyxl import load_workbook , Workbook
from tkinter import *
from datetime import date,datetime
import os
import Print_File

wb_c = Workbook()
folder_name = str(date.today().strftime("%b"))+'-'+str(datetime.now().date().year)
file_name = 'Appointment_Token_List '+date.today().strftime("%b-%d-%Y")+'.xlsx'

if not os.path.isdir(folder_name):
    os.mkdir(folder_name)

if os.path.isfile(os.path.join(folder_name,file_name)):
    # opening the existing excel file
    wb = load_workbook(os.path.join(folder_name,file_name))
    # create the sheet object
    sheet = wb.active

else:
    ws = wb_c.active
    ws.title = "Today's Sheet"
    wb_c.save(filename=os.path.join(folder_name,file_name))
    # opening the existing excel file
    wb = load_workbook(os.path.join(folder_name,file_name))
    # create the sheet object
    sheet = wb.active

def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Token"
    sheet.cell(row=1, column=3).value = "Doctor"
    sheet.cell(row=1, column=4).value = "Date"
    sheet.cell(row=1, column=5).value = "Patient ID"
    sheet.cell(row=1, column=6).value = "Consultation Charge"
    sheet.cell(row=1, column=7).value = "Address"



# Function to set focus (cursor)
def focus1(event):
    # set focus on the token_field box
    token_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the doctor box
    doctor_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the date_field box
    date_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the patient_id_no_field box
    patient_id_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the consultation_charge_field box
    consultation_charge_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    token_field.delete(0, END)
    doctor_field.delete(0, END)
    date_field.delete(0, END)
    patient_id_field.delete(0, END)
    consultation_charge_field.delete(0, END)
    address_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" and
            token_field.get() == "" and
            doctor_field.get() == "" and
            date_field.get() == "" and
            patient_id_field.get() == "" and
            consultation_charge_field.get() == "" and
            address_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        Print_File.create_text_file(name_field.get(), token_field.get(), date.today().strftime("%b-%d-%Y"),
                                    doctor_field.get(),
                                    patient_id_field.get(), consultation_charge_field.get())

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = token_field.get()
        sheet.cell(row=current_row + 1, column=3).value = doctor_field.get()
        sheet.cell(row=current_row + 1, column=4).value = date_field.get()
        sheet.cell(row=current_row + 1, column=5).value = patient_id_field.get()
        sheet.cell(row=current_row + 1, column=6).value = consultation_charge_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # save the file
        wb.save(os.path.join(folder_name,file_name))

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()



# Driver code
if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("Token Registration Form GUI")

    # set the configuration of GUI window
    root.geometry("650x350")

    excel()

    # create a Date label
    heading = Label(root, text="Token Generation Application Form", bg="light green",font=("Arial", 20))

    # create a patient_id No. label
    patient_id = Label(root, text="Patient ID", bg="light green", font=("Arial", 11))

    # create a Name label
    name = Label(root, text="Patient Name", bg="light green" , font=("Arial", 11))

    # create a token label
    token = Label(root, text="Token", bg="light green", font=("Arial", 11))

    # create a Doctor label
    doctor = Label(root, text="Doctor", bg="light green", font=("Arial", 11))

    # create a Date No. label
    date_no = Label(root, text="Today's Date ", bg="light green", font=("Arial", 11))

    # create a consultation_charge label
    consultation_charge = Label(root, text="Consultation Charge (Rs.)", bg="light green", font=("Arial", 11))

    # create a address label
    address = Label(root, text="Address", bg="light green", font=("Arial", 11))

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    token.grid(row=2, column=0)
    doctor.grid(row=3, column=0)
    date_no.grid(row=4, column=0)
    patient_id.grid(row=5, column=0)
    consultation_charge.grid(row=6, column=0)
    address.grid(row=7, column=0)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    token_field = Entry(root)
    doctor_field = Entry(root)
    date_field = Entry(root)
    patient_id_field = Entry(root)
    consultation_charge_field = Entry(root)
    address_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    token_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    doctor_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    date_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    patient_id_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    consultation_charge_field.bind("<Return>", focus6)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="120")
    token_field.grid(row=2, column=1, ipadx="120")
    doctor_field.grid(row=3, column=1, ipadx="120")
    date_field.grid(row=4, column=1, ipadx="120")
    patient_id_field.grid(row=5, column=1, ipadx="120")
    consultation_charge_field.grid(row=6, column=1, ipadx="120")
    address_field.grid(row=7, column=1, ipadx="120")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Create Token", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()
